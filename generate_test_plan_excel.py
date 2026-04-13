#!/usr/bin/env python3
"""Generate Excel test plan from the markdown test plan data."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Styles
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
section_font = Font(bold=True, size=11, color="FFFFFF")
section_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
subsection_font = Font(bold=True, size=10, color="2F5496")
subsection_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
blocker_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
important_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
nice_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap = Alignment(wrap_text=True, vertical='top')

def style_header_row(ws, row, cols=7):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def add_section_row(ws, row, title, cols=7):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = Alignment(horizontal='left', vertical='center')
    for c in range(1, cols+1):
        ws.cell(row=row, column=c).border = thin_border
        ws.cell(row=row, column=c).fill = section_fill

def add_subsection_row(ws, row, title, cols=7):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = subsection_font
    cell.fill = subsection_fill
    cell.alignment = Alignment(horizontal='left', vertical='center')
    for c in range(1, cols+1):
        ws.cell(row=row, column=c).border = thin_border
        ws.cell(row=row, column=c).fill = subsection_fill

def add_test_row(ws, row, test_id, test_desc, priority, day, section, notes=""):
    prio_map = {"Blocker": blocker_fill, "Important": important_fill, "Nice-to-have": nice_fill}
    ws.cell(row=row, column=1, value=test_id).border = thin_border
    ws.cell(row=row, column=2, value=day).border = thin_border
    ws.cell(row=row, column=3, value=section).border = thin_border
    ws.cell(row=row, column=4, value=test_desc).border = thin_border
    ws.cell(row=row, column=4).alignment = wrap
    p_cell = ws.cell(row=row, column=5, value=priority)
    p_cell.border = thin_border
    if priority in prio_map:
        p_cell.fill = prio_map[priority]
    p_cell.alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=6, value="").border = thin_border  # Result
    ws.cell(row=row, column=6).alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=7, value=notes).border = thin_border
    ws.cell(row=row, column=7).alignment = wrap

# ========== SHEET 1: ALL TESTS ==========
ws = wb.active
ws.title = "Test Plan"
ws.sheet_properties.tabColor = "2F5496"

# Column widths
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 60
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 45

# Title row
ws.merge_cells('A1:G1')
title_cell = ws.cell(row=1, column=1, value="Proper Place — Pre-Release Test Plan (v2.2.0)")
title_cell.font = Font(bold=True, size=14, color="2F5496")
title_cell.alignment = Alignment(horizontal='center')

ws.merge_cells('A2:G2')
ws.cell(row=2, column=1, value="Tester: Pierce Shapton | Platforms: iOS, Android, Web | Date: _______________").alignment = Alignment(horizontal='center')

# Header row
row = 4
headers = ["Test #", "Day", "Section", "Test Description", "Priority", "Result", "Notes"]
for i, h in enumerate(headers, 1):
    ws.cell(row=row, column=i, value=h)
style_header_row(ws, row)

row = 5

# All test data organized by section
tests = [
    # DAY 1
    ("DAY 1 — CORE USER FLOWS & PAYMENTS", None),
    
    ("1. AUTHENTICATION", "sub"),
    ("1.1 Fresh Install & Welcome", "subsub"),
    ("1.1.1", "App opens to Welcome Screen with background image and branding", "Important", "1", "Authentication"),
    ("1.1.2", '"Login" button navigates to Login Screen', "Blocker", "1", "Authentication"),
    ("1.1.3", '"Sign Up" button navigates to Signup Screen', "Blocker", "1", "Authentication"),
    ("1.1.4", "App branding/logo displays correctly (no broken images)", "Important", "1", "Authentication"),
    
    ("1.2 Sign Up", "subsub"),
    ("1.2.1", "Enter valid name, email, password, confirm password", "Blocker", "1", "Authentication"),
    ("1.2.2", "Vehicle registration field accepts valid UK plates (e.g. AB12 CDE)", "Important", "1", "Authentication"),
    ("1.2.3", "Vehicle dimension sliders work (height/width/length) and save", "Important", "1", "Authentication"),
    ("1.2.4", "Password mismatch shows warning message", "Blocker", "1", "Authentication"),
    ("1.2.5", "Password visibility toggle (show/hide) works on both fields", "Nice-to-have", "1", "Authentication"),
    ("1.2.6", "Weak password (less than 6 chars) shows error", "Blocker", "1", "Authentication"),
    ("1.2.7", "Duplicate email shows appropriate error", "Blocker", "1", "Authentication"),
    ("1.2.8", "Empty required fields show validation errors", "Blocker", "1", "Authentication"),
    ("1.2.9", "Successful signup navigates to Email Verification Screen", "Blocker", "1", "Authentication"),
    ("1.2.10", "Rate limiting: 6th signup attempt within an hour is blocked", "Important", "1", "Authentication"),
    
    ("1.3 Email Verification", "subsub"),
    ("1.3.1", "Email Verification Screen displays with correct email address", "Blocker", "1", "Authentication"),
    ("1.3.2", "Verification email arrives in inbox (check spam)", "Blocker", "1", "Authentication", "⚠️ SMTP may not be working in production yet"),
    ("1.3.3", "Clicking verification link in email marks user as verified", "Blocker", "1", "Authentication"),
    ("1.3.4", "App auto-detects verification (polls every 5s) and proceeds to home", "Blocker", "1", "Authentication"),
    ("1.3.5", '"Resend" button sends a new verification email', "Important", "1", "Authentication"),
    ("1.3.6", "Back button returns to Welcome Screen", "Important", "1", "Authentication"),
    ("1.3.7", "Logging in with unverified account shows verification screen", "Blocker", "1", "Authentication"),
    
    ("1.4 Login", "subsub"),
    ("1.4.1", "Valid email + password logs in successfully", "Blocker", "1", "Authentication"),
    ("1.4.2", "Invalid password shows error message", "Blocker", "1", "Authentication"),
    ("1.4.3", "Non-existent email shows error (shouldn't reveal if account exists)", "Blocker", "1", "Authentication"),
    ("1.4.4", '"Remember Me" checkbox persists login across app restart', "Important", "1", "Authentication"),
    ("1.4.5", "Password visibility toggle works", "Nice-to-have", "1", "Authentication"),
    ("1.4.6", "Rate limiting: 6th login attempt within 15 mins is blocked", "Important", "1", "Authentication"),
    
    ("1.5 Forgot Password", "subsub"),
    ("1.5.1", '"Forgot Password" link from login screen works', "Blocker", "1", "Authentication"),
    ("1.5.2", "Entering valid email shows success message", "Blocker", "1", "Authentication"),
    ("1.5.3", "Reset email arrives with valid link", "Blocker", "1", "Authentication", "⚠️ Depends on SMTP fix"),
    ("1.5.4", "Reset link opens password reset form", "Blocker", "1", "Authentication"),
    ("1.5.5", "Setting new password works and allows login", "Blocker", "1", "Authentication"),
    ("1.5.6", "Entering non-existent email still shows success (no info leak)", "Important", "1", "Authentication"),
    
    ("1.6 Session Management", "subsub"),
    ("1.6.1", "Token auto-refreshes (stay logged in for 30+ mins without action)", "Blocker", "1", "Authentication"),
    ("1.6.2", "Logout clears session and returns to Welcome Screen", "Blocker", "1", "Authentication"),
    ("1.6.3", "After logout, pressing back doesn't return to authenticated screens", "Important", "1", "Authentication"),
    ("1.6.4", "Force-closing and reopening app maintains session (if Remember Me)", "Important", "1", "Authentication"),
    
    ("2. MAP & BROWSING", "sub"),
    ("2.1 Map Display", "subsub"),
    ("2.1.1", "Map loads with Google Maps tiles", "Blocker", "1", "Map & Browsing"),
    ("2.1.2", "Current location button works (if permissions granted)", "Important", "1", "Map & Browsing"),
    ("2.1.3", 'Custom "P" pin markers appear when zoomed in (zoom 11+)', "Blocker", "1", "Map & Browsing"),
    ("2.1.4", "Markers disappear when zoomed out too far", "Nice-to-have", "1", "Map & Browsing"),
    ("2.1.5", "Pin tip points at the exact location (not offset)", "Important", "1", "Map & Browsing"),
    ("2.1.6", "Map type toggle works (satellite/standard)", "Nice-to-have", "1", "Map & Browsing"),
    ("2.1.7", "Map remembers last position between sessions", "Nice-to-have", "1", "Map & Browsing"),
    
    ("2.2 Place Discovery", "subsub"),
    ("2.2.1", "Tapping a marker shows place preview card", "Blocker", "1", "Map & Browsing"),
    ("2.2.2", "Place preview shows: name, price, image thumbnail", "Blocker", "1", "Map & Browsing"),
    ("2.2.3", "Tapping preview navigates to Place Detail Screen", "Blocker", "1", "Map & Browsing"),
    ("2.2.4", "Facility filter buttons appear and are functional", "Important", "1", "Map & Browsing"),
    ("2.2.5", "Filtering by facility updates visible markers", "Important", "1", "Map & Browsing"),
    ("2.2.6", "Search functionality finds places by name/location", "Important", "1", "Map & Browsing"),
    ("2.2.7", "Vehicle size filter toggle filters out places too small for your van", "Important", "1", "Map & Browsing"),
    
    ("2.3 Route Planning", "subsub"),
    ("2.3.1", "Route planning (start/destination) draws a route on map", "Important", "1", "Map & Browsing"),
    ("2.3.2", "Search radius circle displays around destination", "Nice-to-have", "1", "Map & Browsing"),
    
    ("2.4 Offline Maps", "subsub"),
    ("2.4.1", "Offline regions screen lists all 16 UK regions", "Important", "1", "Map & Browsing"),
    ("2.4.2", "Downloading a region shows progress indicator", "Important", "1", "Map & Browsing"),
    ("2.4.3", "Downloaded region data persists when offline", "Important", "1", "Map & Browsing"),
    ("2.4.4", "Cache size displays correctly", "Nice-to-have", "1", "Map & Browsing"),
    
    ("3. PLACE DETAIL", "sub"),
    ("3.1", "Place Detail Screen loads with all information", "Blocker", "1", "Place Detail"),
    ("3.2", "Image carousel swipes between photos", "Blocker", "1", "Place Detail"),
    ("3.3", "Place name, description, address display correctly", "Blocker", "1", "Place Detail"),
    ("3.4", "Price per night displays correctly (£)", "Blocker", "1", "Place Detail"),
    ("3.5", "Amenities list shows correct icons", "Important", "1", "Place Detail"),
    ("3.6", "Reviews section shows existing reviews with star ratings", "Important", "1", "Place Detail"),
    ("3.7", "Average rating calculates correctly", "Important", "1", "Place Detail"),
    ("3.8", "Review photos display in reviews", "Nice-to-have", "1", "Place Detail"),
    ("3.9", "Availability calendar shows booked dates in green", "Blocker", "1", "Place Detail"),
    ("3.10", "Vehicle fit warning appears if van is too large", "Important", "1", "Place Detail"),
    ("3.11", "Business description, food menu, access route show if available", "Nice-to-have", "1", "Place Detail"),
    ("3.12", "Host info (name) displays", "Important", "1", "Place Detail"),
    ("3.13", '"Add to Favourites" heart icon works', "Important", "1", "Place Detail"),
    ("3.14", "Place is marked as favourite and appears in Favourites screen", "Important", "1", "Place Detail"),
    
    ("4. BOOKING FLOW & PAYMENTS", "sub"),
    ("4.1 Creating a Booking", "subsub"),
    ("4.1.1", '"Book Now" button opens booking calendar', "Blocker", "1", "Booking & Payment"),
    ("4.1.2", "Select check-in date on calendar (orange highlight)", "Blocker", "1", "Booking & Payment"),
    ("4.1.3", "Select check-out date (range highlighted)", "Blocker", "1", "Booking & Payment"),
    ("4.1.4", "Cannot select dates that are already booked (green)", "Blocker", "1", "Booking & Payment"),
    ("4.1.5", "Cannot select past dates", "Blocker", "1", "Booking & Payment"),
    ("4.1.6", "Check-in/check-out times can be adjusted", "Important", "1", "Booking & Payment"),
    ("4.1.7", "Early arrival fee (£5/hr before standard time) calculates", "Important", "1", "Booking & Payment"),
    ("4.1.8", "Late departure fee (£5/hr after standard time) calculates", "Important", "1", "Booking & Payment"),
    ("4.1.9", "Total price displays correctly (nights × price + any fees)", "Blocker", "1", "Booking & Payment"),
    ("4.1.10", "Van registration is pre-filled from profile or can be entered", "Important", "1", "Booking & Payment"),
    ("4.1.11", "UK number plate validation works (rejects invalid formats)", "Important", "1", "Booking & Payment"),
    ("4.1.12", "Capacity check: booking rejected if place is full", "Blocker", "1", "Booking & Payment"),
    
    ("4.2 Stripe Payment — Test Cards: 4242424242424242 (success), 4000000000000002 (decline), 4000002500003155 (3DS)", "subsub"),
    ("4.2.1", "Stripe payment sheet appears after confirming booking details", "Blocker", "1", "Booking & Payment"),
    ("4.2.2", "Successful payment with test card 4242... creates booking", "Blocker", "1", "Booking & Payment"),
    ("4.2.3", 'Booking status shows as "Pending" (awaiting host approval)', "Blocker", "1", "Booking & Payment"),
    ("4.2.4", "Declined card (4000...0002) shows appropriate error message", "Blocker", "1", "Booking & Payment"),
    ("4.2.5", "3D Secure card (4000...3155) shows authentication challenge", "Important", "1", "Booking & Payment"),
    ("4.2.6", "Payment amount matches displayed total", "Blocker", "1", "Booking & Payment"),
    ("4.2.7", "Funds are HELD (not captured) — check Stripe Dashboard", "Blocker", "1", "Booking & Payment"),
    ("4.2.8", "Booking confirmation screen displays after successful payment", "Blocker", "1", "Booking & Payment"),
    ("4.2.9", "Booking reference number generated (PP-YYMMDD-XXXX format)", "Important", "1", "Booking & Payment"),
    
    ("4.3 Booking Management (Guest)", "subsub"),
    ("4.3.1", 'New booking appears in "My Bookings" screen', "Blocker", "1", "Booking & Payment"),
    ("4.3.2", "Booking tabs filter correctly (All/Pending/Confirmed/Completed)", "Important", "1", "Booking & Payment"),
    ("4.3.3", "Calendar view toggle shows bookings on calendar", "Important", "1", "Booking & Payment"),
    ("4.3.4", "Search within bookings works", "Nice-to-have", "1", "Booking & Payment"),
    ("4.3.5", "Tapping a booking opens Booking Detail Screen", "Blocker", "1", "Booking & Payment"),
    ("4.3.6", "Booking detail shows: dates, place, status, total paid", "Blocker", "1", "Booking & Payment"),
    ("4.3.7", '"Get Directions" button opens Waze/Maps', "Important", "1", "Booking & Payment"),
    ("4.3.8", '"Chat with Host" button opens messaging', "Blocker", "1", "Booking & Payment"),
    ("4.3.9", "Cancel booking releases the payment hold", "Blocker", "1", "Booking & Payment"),
    ("4.3.10", "Cancelled booking status updates correctly", "Blocker", "1", "Booking & Payment"),
    
    ("5. MESSAGING SYSTEM", "sub"),
    ("5.1 Guest → Host Chat", "subsub"),
    ("5.1.1", "Opening chat from booking shows conversation", "Blocker", "1", "Messaging"),
    ("5.1.2", "Sending a text message works", "Blocker", "1", "Messaging"),
    ("5.1.3", "Messages appear in correct order (newest at bottom)", "Blocker", "1", "Messaging"),
    ("5.1.4", "Message polling (3s) picks up new messages from host", "Blocker", "1", "Messaging"),
    ("5.1.5", "Unread message badge shows on Bookings/Messages tab", "Blocker", "1", "Messaging"),
    ("5.1.6", "Badge count clears when conversation is opened", "Important", "1", "Messaging"),
    ("5.1.7", "All Chats screen lists all conversations", "Blocker", "1", "Messaging"),
    ("5.1.8", "Swipe actions: mark unread, delete work", "Nice-to-have", "1", "Messaging"),
    
    ("5.2 Chat Lifecycle", "subsub"),
    ("5.2.1", 'Chat is "open" during active booking', "Important", "1", "Messaging"),
    ("5.2.2", 'Chat shows "closing soon" near end of stay', "Important", "1", "Messaging"),
    ("5.2.3", "Chat closes after booking completion", "Important", "1", "Messaging"),
    ("5.2.4", '"Request Reopen" button available on closed chats', "Important", "1", "Messaging"),
    ("5.2.5", "Host can approve/deny reopen request", "Important", "1", "Messaging"),
    ("5.3.1", 'Host response time label displays (e.g. "Usually responds within 1 hour")', "Nice-to-have", "1", "Messaging"),
    
    ("6. REVIEWS", "sub"),
    ("6.1", 'After a completed booking, "Leave Review" option appears', "Blocker", "1", "Reviews"),
    ("6.2", "Star rating selector works (1-5 stars)", "Blocker", "1", "Reviews"),
    ("6.3", "Text comment field accepts input", "Blocker", "1", "Reviews"),
    ("6.4", "Photo upload (up to 5 photos) works", "Important", "1", "Reviews"),
    ("6.5", "EXIF metadata stripped from photos (privacy)", "Important", "1", "Reviews", "Upload photo with GPS, verify stripped"),
    ("6.6", "Submitting review succeeds", "Blocker", "1", "Reviews"),
    ("6.7", "Review appears on the place's detail page", "Blocker", "1", "Reviews"),
    ("6.8", "Average rating updates after new review", "Important", "1", "Reviews"),
    
    ("7. FAVOURITES", "sub"),
    ("7.1", "Adding a place to favourites from detail screen works", "Important", "1", "Favourites"),
    ("7.2", "Favourites tab shows all saved places", "Important", "1", "Favourites"),
    ("7.3", "Removing a favourite updates the list", "Important", "1", "Favourites"),
    ("7.4", "Favourite heart markers appear on map", "Nice-to-have", "1", "Favourites"),
    ("7.5", "Search/filter within favourites works", "Nice-to-have", "1", "Favourites"),
    
    # DAY 2
    ("DAY 2 — HOST, ADMIN, WEBSITE & EDGE CASES", None),
    
    ("8. HOST APPLICATION & ONBOARDING", "sub"),
    ("8.1 Become a Host", "subsub"),
    ("8.1.1", '"Become a Host" option in More menu visible for regular users', "Blocker", "2", "Host Onboarding"),
    ("8.1.2", "Host application form loads with all fields", "Blocker", "2", "Host Onboarding"),
    ("8.1.3", "Contact name, email, phone fields work", "Blocker", "2", "Host Onboarding"),
    ("8.1.4", "Business description field works", "Important", "2", "Host Onboarding"),
    ("8.1.5", "Address field uses Google Places autocomplete", "Blocker", "2", "Host Onboarding"),
    ("8.1.6", "Business type selector works (pub, farm, etc.)", "Important", "2", "Host Onboarding"),
    ("8.1.7", "Number of van spaces field works", "Important", "2", "Host Onboarding"),
    ("8.1.8", "Referral code field works (optional)", "Nice-to-have", "2", "Host Onboarding"),
    ("8.1.9", "GPS location auto-populates from address", "Important", "2", "Host Onboarding"),
    ("8.1.10", "Submitting application shows success message", "Blocker", "2", "Host Onboarding"),
    
    ("8.2 Host Contract", "subsub"),
    ("8.2.1", "After application approval, Host Contract screen appears", "Blocker", "2", "Host Onboarding"),
    ("8.2.2", "Contract text is readable and scrollable", "Blocker", "2", "Host Onboarding"),
    ("8.2.3", 'Must scroll to bottom before "Accept" becomes enabled', "Blocker", "2", "Host Onboarding"),
    ("8.2.4", "Accepting contract records acceptance (IP, timestamp)", "Blocker", "2", "Host Onboarding"),
    ("8.2.5", "After contract, Stripe Payout Setup screen appears", "Blocker", "2", "Host Onboarding"),
    
    ("8.3 Stripe Connect (Host Payouts)", "subsub"),
    ("8.3.1", "Stripe Connect onboarding opens external browser", "Blocker", "2", "Host Onboarding"),
    ("8.3.2", "Completing Stripe onboarding returns to app", "Blocker", "2", "Host Onboarding"),
    ("8.3.3", "App verifies Stripe Connect status on return", "Blocker", "2", "Host Onboarding"),
    ("8.3.4", 'If onboarding incomplete, shows "Complete Setup" option', "Important", "2", "Host Onboarding"),
    ("8.3.5", "Payout status visible on Host Dashboard", "Important", "2", "Host Onboarding"),
    
    ("9. HOST SITE MANAGEMENT", "sub"),
    ("9.1 Create a Site", "subsub"),
    ("9.1.1", '"Create Site" button available in Host mode', "Blocker", "2", "Host Sites"),
    ("9.1.2", "All form fields work: name, address, description", "Blocker", "2", "Host Sites"),
    ("9.1.3", "Access route description field works", "Important", "2", "Host Sites"),
    ("9.1.4", "Price per night field (numeric input) works", "Blocker", "2", "Host Sites"),
    ("9.1.5", "Website URL field works", "Nice-to-have", "2", "Host Sites"),
    ("9.1.6", "Business name and description fields work", "Important", "2", "Host Sites"),
    ("9.1.7", "Food menu description field works", "Nice-to-have", "2", "Host Sites"),
    ("9.1.8", "Main photo upload works", "Blocker", "2", "Host Sites"),
    ("9.1.9", "Supporting photos upload (multiple) works", "Important", "2", "Host Sites"),
    ("9.1.10", "Business photos upload works", "Nice-to-have", "2", "Host Sites"),
    ("9.1.11", "Vehicle dimension limits with metric/imperial toggle", "Important", "2", "Host Sites"),
    ("9.1.12", "Number of van spaces field works", "Blocker", "2", "Host Sites"),
    ("9.1.13", "Google Maps pin placement for exact location", "Blocker", "2", "Host Sites"),
    ("9.1.14", 'Saving site shows success and appears in "My Sites"', "Blocker", "2", "Host Sites"),
    ("9.1.15", 'New site status is "Pending" (awaiting admin approval)', "Blocker", "2", "Host Sites"),
    
    ("9.2 Edit a Site", "subsub"),
    ("9.2.1", "Tapping a site opens it for editing", "Blocker", "2", "Host Sites"),
    ("9.2.2", "All fields are pre-populated with existing data", "Blocker", "2", "Host Sites"),
    ("9.2.3", "Changes save correctly", "Blocker", "2", "Host Sites"),
    ("9.2.4", "Adding/removing photos works", "Important", "2", "Host Sites"),
    
    ("9.3 My Sites List", "subsub"),
    ("9.3.1", "Sites list shows all host's sites", "Blocker", "2", "Host Sites"),
    ("9.3.2", "Status badges display correctly (Drafting/Pending/Approved/Rejected)", "Blocker", "2", "Host Sites"),
    ("9.3.3", "Only approved sites appear on the public map", "Blocker", "2", "Host Sites"),
    ("9.3.4", "Mark site as unavailable toggles it off the map", "Important", "2", "Host Sites"),
    ("9.3.5", "Mark site as available restores it", "Important", "2", "Host Sites"),
    
    ("10. HOST BOOKING MANAGEMENT", "sub"),
    ("10.1", "Host sees incoming bookings in Host Bookings screen", "Blocker", "2", "Host Bookings"),
    ("10.2", "Booking shows guest info, dates, van details", "Blocker", "2", "Host Bookings"),
    ("10.3", '"Approve" button captures the held payment (check Stripe)', "Blocker", "2", "Host Bookings"),
    ("10.4", 'After approval, booking status changes to "Confirmed"', "Blocker", "2", "Host Bookings"),
    ("10.5", '"Reject" button cancels the payment hold (check Stripe)', "Blocker", "2", "Host Bookings"),
    ("10.6", 'After rejection, booking status changes to "Rejected"', "Blocker", "2", "Host Bookings"),
    ("10.7", "Calendar view shows bookings on dates", "Important", "2", "Host Bookings"),
    ("10.8", "Filter tabs work (Confirmed/Pending/Completed/All)", "Important", "2", "Host Bookings"),
    ("10.9", "Unread message badges appear per booking", "Important", "2", "Host Bookings"),
    ("10.10", "15% platform fee deducted from host payout (check Stripe)", "Blocker", "2", "Host Bookings"),
    
    ("11. HOST DASHBOARD & FEATURES", "sub"),
    ("11.1", "Dashboard shows: unread messages, total bookings, revenue", "Blocker", "2", "Host Dashboard"),
    ("11.2", "Pending payments count is accurate", "Important", "2", "Host Dashboard"),
    ("11.3", "Average length of stay calculates", "Nice-to-have", "2", "Host Dashboard"),
    ("11.4", "Recent conversations list is clickable", "Important", "2", "Host Dashboard"),
    ("11.5", "Recent bookings list is clickable", "Important", "2", "Host Dashboard"),
    
    ("11.1 Auto-Messages", "subsub"),
    ("11.1.1", "Auto-message config screen accessible per place", "Important", "2", "Host Dashboard"),
    ("11.1.2", 'Can set message for "On Booking" trigger', "Important", "2", "Host Dashboard"),
    ("11.1.3", 'Can set message for "24h Before Check-in" trigger', "Important", "2", "Host Dashboard"),
    ("11.1.4", 'Can set message for "1h Before Arrival" trigger', "Important", "2", "Host Dashboard"),
    ("11.1.5", 'Can set message for "At Checkout" trigger', "Important", "2", "Host Dashboard"),
    ("11.1.6", "Auto-message actually sends at the correct trigger time", "Important", "2", "Host Dashboard", "Hard to test — check DB or wait"),
    
    ("11.2 Host Reviews & Chat", "subsub"),
    ("11.2.1", "Host Reviews screen shows all reviews across all places", "Important", "2", "Host Dashboard"),
    ("11.2.2", "Filter by star rating works", "Nice-to-have", "2", "Host Dashboard"),
    ("11.2.3", "Review photos display correctly", "Nice-to-have", "2", "Host Dashboard"),
    ("11.3.1", "Host chat screen lists all guest conversations", "Blocker", "2", "Host Dashboard"),
    ("11.3.2", "Open/closed sections display correctly", "Important", "2", "Host Dashboard"),
    ("11.3.3", "Sending messages to guests works", "Blocker", "2", "Host Dashboard"),
    ("11.3.4", "Reopen request handling works (approve/deny)", "Important", "2", "Host Dashboard"),
    ("11.4.1", "Host can rate a guest after completed booking", "Important", "2", "Host Dashboard"),
    ("11.4.2", "Guest rating is stored and retrievable", "Important", "2", "Host Dashboard"),
    
    ("12. ADMIN PANEL — Login: admin@properplace.com / AdminPass123!", "sub"),
    ("12.1.1", "Admin Dashboard loads with quick-action cards", "Blocker", "2", "Admin Panel"),
    ("12.1.2", "Booking search works", "Important", "2", "Admin Panel"),
    ("12.2.1", "Pending/Approved/Rejected tabs show correct counts", "Blocker", "2", "Admin Panel"),
    ("12.2.2", "Tapping a pending place shows full detail with images", "Blocker", "2", "Admin Panel"),
    ("12.2.3", "Host info displayed (total sites, join date, contract status)", "Important", "2", "Admin Panel"),
    ("12.2.4", "Google Maps preview shows place location", "Important", "2", "Admin Panel"),
    ("12.2.5", '"Approve" button approves and place appears on map', "Blocker", "2", "Admin Panel"),
    ("12.2.6", '"Reject" button rejects with reason', "Blocker", "2", "Admin Panel"),
    ("12.2.7", '"Remove" button removes an approved place', "Important", "2", "Admin Panel"),
    ("12.3.1", "All platform bookings are visible", "Blocker", "2", "Admin Panel"),
    ("12.3.2", "Filter by status works (All/Pending/Confirmed/Completed/Cancelled)", "Important", "2", "Admin Panel"),
    ("12.4.1", "Admin can view all conversations", "Important", "2", "Admin Panel"),
    ("12.4.2", "Admin can send messages", "Important", "2", "Admin Panel"),
    ("12.5.1", "Contact submissions appear with status (new/read/responded/closed)", "Blocker", "2", "Admin Panel"),
    ("12.5.2", "Filter by status works", "Important", "2", "Admin Panel"),
    ("12.5.3", "Filter by category (hosts/users) works", "Important", "2", "Admin Panel"),
    ("12.5.4", "Can update contact status", "Important", "2", "Admin Panel"),
    
    ("13. PROFILE & SETTINGS", "sub"),
    ("13.1", "Profile screen shows current name, email, phone, bio", "Blocker", "2", "Profile"),
    ("13.2", "Editing name saves correctly", "Blocker", "2", "Profile"),
    ("13.3", "Editing phone number saves correctly", "Important", "2", "Profile"),
    ("13.4", "Editing bio saves correctly", "Nice-to-have", "2", "Profile"),
    ("13.5", "Vehicle registration can be updated", "Important", "2", "Profile"),
    ("13.6", "Vehicle dimensions screen works (ft/m toggle, sliders)", "Important", "2", "Profile"),
    ("13.7", '"Contact Us" form accessible and submits', "Important", "2", "Profile"),
    ("13.8", "T&C and Privacy links open correctly", "Important", "2", "Profile"),
    ("13.9", "Logout works from More menu", "Blocker", "2", "Profile"),
    
    ("14. PUSH NOTIFICATIONS", "sub"),
    ("14.1", "App requests notification permission on first launch", "Important", "2", "Notifications"),
    ("14.2", "Device token registered with backend", "Important", "2", "Notifications"),
    ("14.3", "Push notification received for new booking (as host)", "Important", "2", "Notifications"),
    ("14.4", "Push notification received for new message", "Important", "2", "Notifications"),
    ("14.5", "Tapping notification opens relevant screen", "Important", "2", "Notifications"),
    ("14.6", "Badge counts update on bottom nav bar", "Blocker", "2", "Notifications"),
    ("14.7", "Device token unregistered on logout", "Nice-to-have", "2", "Notifications"),
    
    ("15. WEBSITE (proper-place.co.uk)", "sub"),
    ("15.1 Public Pages", "subsub"),
    ("15.1.1", "Homepage loads with hero, featured places, how-it-works", "Blocker", "2", "Website"),
    ("15.1.2", "Navigation bar links all work", "Blocker", "2", "Website"),
    ("15.1.3", "Footer links all work", "Important", "2", "Website"),
    ("15.1.4", "Cookie consent banner appears and works", "Blocker", "2", "Website"),
    ("15.1.5", "/browse — Google Maps loads with place markers", "Blocker", "2", "Website"),
    ("15.1.6", "/browse — Clicking marker shows place card", "Blocker", "2", "Website"),
    ("15.1.7", "/place/[id] — Place detail page loads with images, amenities, reviews", "Blocker", "2", "Website"),
    ("15.1.8", "/about — About page loads", "Important", "2", "Website"),
    ("15.1.9", "/how-it-works — Page loads with steps for guests and hosts", "Important", "2", "Website"),
    ("15.1.10", "/download — App download page with store buttons", "Important", "2", "Website"),
    ("15.1.11", "/contact — Contact form submits successfully", "Blocker", "2", "Website"),
    ("15.1.12", "/privacy — Privacy policy loads", "Blocker", "2", "Website"),
    ("15.1.13", "/terms — Terms of service loads", "Blocker", "2", "Website"),
    ("15.1.14", "/cookies — Cookie policy loads", "Important", "2", "Website"),
    ("15.1.15", "/scan — QR code landing page loads", "Nice-to-have", "2", "Website"),
    ("15.1.16", "/become-host — Host recruitment page loads", "Important", "2", "Website"),
    ("15.1.17", "/host-signup — Host lead form submits", "Important", "2", "Website"),
    ("15.1.18", "/qr-codes — QR code generator works, PNG download works", "Nice-to-have", "2", "Website"),
    
    ("15.2 Website Auth", "subsub"),
    ("15.2.1", "/auth/signup — Registration works with all fields", "Blocker", "2", "Website"),
    ("15.2.2", "Referral code field on signup works", "Important", "2", "Website"),
    ("15.2.3", "/auth/login — Login works", "Blocker", "2", "Website"),
    ("15.2.4", "Forgot password flow works from login page", "Blocker", "2", "Website"),
    ("15.2.5", "Token auto-refresh works (stay logged in)", "Important", "2", "Website"),
    
    ("15.3 Website Booking Flow", "subsub"),
    ("15.3.1", "/place/[id]/book — Booking form loads", "Blocker", "2", "Website"),
    ("15.3.2", "Date picker works with unavailable dates blocked", "Blocker", "2", "Website"),
    ("15.3.3", "Vehicle registration and phone fields work", "Important", "2", "Website"),
    ("15.3.4", "Vehicle dimensions can be entered", "Important", "2", "Website"),
    ("15.3.5", "Price calculation displays correctly", "Blocker", "2", "Website"),
    ("15.3.6", "Stripe payment works (test card)", "Blocker", "2", "Website"),
    ("15.3.7", "Booking success confirmation shows", "Blocker", "2", "Website"),
    
    ("15.4 Website Dashboard", "subsub"),
    ("15.4.1", "/dashboard — Overview loads with role-based content", "Blocker", "2", "Website"),
    ("15.4.2", "/dashboard/bookings — Guest bookings list with filters", "Blocker", "2", "Website"),
    ("15.4.3", "/dashboard/bookings/[id] — Booking detail, cancel, review", "Blocker", "2", "Website"),
    ("15.4.4", "/dashboard/messages — Conversations list", "Blocker", "2", "Website"),
    ("15.4.5", "/dashboard/messages/[userId] — Chat works with polling", "Blocker", "2", "Website"),
    ("15.4.6", "/dashboard/profile — Profile, Vehicle, Security tabs", "Important", "2", "Website"),
    ("15.4.7", "/dashboard/referrals — Referral code, stats, share", "Important", "2", "Website"),
    ("15.4.8", "/dashboard/places — Host places list (if host)", "Important", "2", "Website"),
    ("15.4.9", "/dashboard/places/new — Create place form", "Important", "2", "Website"),
    ("15.4.10", "/dashboard/host/bookings — Host booking management", "Important", "2", "Website"),
    ("15.4.11", "/dashboard/host/auto-messages — Auto-message config", "Important", "2", "Website"),
    
    ("15.5 Website Admin", "subsub"),
    ("15.5.1", "/dashboard/admin — Admin stats display", "Blocker", "2", "Website"),
    ("15.5.2", "/dashboard/admin/users — User management, role changes", "Important", "2", "Website"),
    ("15.5.3", "/dashboard/admin/places — Place approvals", "Blocker", "2", "Website"),
    ("15.5.4", "/dashboard/admin/bookings — All bookings search", "Important", "2", "Website"),
    ("15.5.5", "/dashboard/admin/contacts — Support ticket management", "Important", "2", "Website"),
    
    ("15.6 Mobile Responsiveness", "subsub"),
    ("15.6.1", "Homepage looks correct on mobile (Chrome DevTools)", "Blocker", "2", "Website"),
    ("15.6.2", "Navigation hamburger menu works on mobile", "Blocker", "2", "Website"),
    ("15.6.3", "Browse map is usable on mobile", "Important", "2", "Website"),
    ("15.6.4", "Booking form works on mobile", "Blocker", "2", "Website"),
    ("15.6.5", "Dashboard sidebar collapses on mobile", "Important", "2", "Website"),
    
    ("16. END-TO-END FLOW TEST (MOST IMPORTANT)", "sub"),
    ("16.1", "Create new guest account (app)", "Blocker", "2", "E2E Flow"),
    ("16.2", "Verify email", "Blocker", "2", "E2E Flow"),
    ("16.3", "Set vehicle dimensions", "Important", "2", "E2E Flow"),
    ("16.4", "Browse map and find a place", "Blocker", "2", "E2E Flow"),
    ("16.5", "View place detail, check amenities and reviews", "Blocker", "2", "E2E Flow"),
    ("16.6", "Add place to favourites", "Important", "2", "E2E Flow"),
    ("16.7", "Start booking: select dates", "Blocker", "2", "E2E Flow"),
    ("16.8", "Complete Stripe payment (test card 4242...)", "Blocker", "2", "E2E Flow"),
    ("16.9", 'Verify booking appears as "Pending"', "Blocker", "2", "E2E Flow"),
    ("16.10", "Send message to host via booking chat", "Blocker", "2", "E2E Flow"),
    ("16.11", "SWITCH TO HOST ACCOUNT", "Blocker", "2", "E2E Flow"),
    ("16.12", "Host sees pending booking in dashboard", "Blocker", "2", "E2E Flow"),
    ("16.13", "Host sees guest message", "Blocker", "2", "E2E Flow"),
    ("16.14", "Host replies to message", "Blocker", "2", "E2E Flow"),
    ("16.15", "Host approves booking", "Blocker", "2", "E2E Flow"),
    ("16.16", "Verify Stripe: payment captured, 15% fee taken", "Blocker", "2", "E2E Flow", "Check Stripe Dashboard"),
    ("16.17", "SWITCH TO GUEST ACCOUNT", "Blocker", "2", "E2E Flow"),
    ("16.18", 'Guest sees booking status changed to "Confirmed"', "Blocker", "2", "E2E Flow"),
    ("16.19", "Guest sees host's reply message", "Blocker", "2", "E2E Flow"),
    ("16.20", "Guest gets directions via Waze/Maps", "Important", "2", "E2E Flow"),
    ("16.21", "After stay: guest leaves a review (stars + text + photo)", "Blocker", "2", "E2E Flow"),
    ("16.22", "SWITCH TO HOST ACCOUNT", "Blocker", "2", "E2E Flow"),
    ("16.23", "Host sees the review", "Blocker", "2", "E2E Flow"),
    ("16.24", "Host rates the guest", "Important", "2", "E2E Flow"),
    ("16.25", "SWITCH TO ADMIN ACCOUNT", "Blocker", "2", "E2E Flow"),
    ("16.26", 'Admin sees the booking in "All Bookings"', "Blocker", "2", "E2E Flow"),
    ("16.27", "Admin sees the review", "Important", "2", "E2E Flow"),
    ("16.28", "Guest makes another booking (same or different place)", "Blocker", "2", "E2E Flow"),
    ("16.29", "Host rejects it", "Blocker", "2", "E2E Flow"),
    ("16.30", 'Guest sees status "Rejected"', "Blocker", "2", "E2E Flow"),
    ("16.31", "Verify Stripe: payment hold released/refunded", "Blocker", "2", "E2E Flow", "Check Stripe Dashboard"),
    ("16.32", "Guest makes a booking and pays", "Blocker", "2", "E2E Flow"),
    ("16.33", "Guest cancels the booking", "Blocker", "2", "E2E Flow"),
    ("16.34", "Verify Stripe: payment hold released", "Blocker", "2", "E2E Flow"),
    ("16.35", 'Booking shows as "Cancelled"', "Blocker", "2", "E2E Flow"),
    
    ("17. HOST SITE SUBMISSION → APPROVAL FLOW", "sub"),
    ("17.1", "Host creates a new site with all details + photos", "Blocker", "2", "Site Approval"),
    ("17.2", 'Site shows as "Pending" in host\'s My Sites', "Blocker", "2", "Site Approval"),
    ("17.3", "LOGIN AS ADMIN", "Blocker", "2", "Site Approval"),
    ("17.4", "Site appears in admin Place Approvals → Pending tab", "Blocker", "2", "Site Approval"),
    ("17.5", "Admin reviews all details (images, host info, map)", "Blocker", "2", "Site Approval"),
    ("17.6", "Admin approves the site", "Blocker", "2", "Site Approval"),
    ("17.7", 'LOGIN AS HOST — site now shows "Approved"', "Blocker", "2", "Site Approval"),
    ("17.8", "LOGIN AS GUEST — site now appears on map", "Blocker", "2", "Site Approval"),
    ("17.9", "Admin rejects a different site with reason", "Important", "2", "Site Approval"),
    ("17.10", 'Host sees "Rejected" status', "Important", "2", "Site Approval"),
    
    ("18. STRIPE FINANCIAL VERIFICATION (Check Stripe Dashboard)", "sub"),
    ("18.1", "PaymentIntent created with capture_method: manual", "Blocker", "2", "Stripe"),
    ("18.2", "On host approval: PaymentIntent status = succeeded", "Blocker", "2", "Stripe"),
    ("18.3", "On host approval: Charge captured", "Blocker", "2", "Stripe"),
    ("18.4", "15% application fee deducted", "Blocker", "2", "Stripe"),
    ("18.5", "On host rejection: PaymentIntent cancelled, no charge", "Blocker", "2", "Stripe"),
    ("18.6", "On guest cancel: Payment hold released", "Blocker", "2", "Stripe"),
    ("18.7", "Host Stripe Connect account shows correct balance", "Blocker", "2", "Stripe"),
    ("18.8", "Webhook events received (check Stripe webhook logs)", "Important", "2", "Stripe"),
    ("18.9", "Expired authorizations auto-cancel (runs hourly)", "Important", "2", "Stripe"),
    
    ("19. CONTACT & SUPPORT", "sub"),
    ("19.1", "App: Contact Us form (More → Contact Us) submits", "Important", "2", "Contact"),
    ("19.2", "App: Category dropdown works", "Important", "2", "Contact"),
    ("19.3", "Website: /contact form submits", "Important", "2", "Contact"),
    ("19.4", "Admin sees contact messages in Admin Contact Messages", "Important", "2", "Contact"),
    ("19.5", "Admin can change status (new → read → responded → closed)", "Important", "2", "Contact"),
    
    ("20. REFERRAL SYSTEM", "sub"),
    ("20.1", "User can view their referral code", "Important", "2", "Referrals"),
    ("20.2", "Referral code can be copied/shared", "Important", "2", "Referrals"),
    ("20.3", "Signing up with a referral code creates a referral record", "Important", "2", "Referrals"),
    ("20.4", "Referral stats display correctly", "Important", "2", "Referrals"),
    ("20.5", "Website: /dashboard/referrals shows stats + code", "Important", "2", "Referrals"),
    
    ("21. EDGE CASES & ERROR HANDLING", "sub"),
    ("21.1", "No internet: app shows offline mode / graceful error", "Important", "2", "Edge Cases"),
    ("21.2", "Slow internet: loading spinners appear", "Important", "2", "Edge Cases"),
    ("21.3", "Back button behaviour on all screens (no crashes)", "Important", "2", "Edge Cases"),
    ("21.4", "Rotate device: layout doesn't break", "Nice-to-have", "2", "Edge Cases"),
    ("21.5", "Very long text in fields (e.g. 1000-char description)", "Nice-to-have", "2", "Edge Cases"),
    ("21.6", "Special characters in name/description (é, ñ, emoji)", "Nice-to-have", "2", "Edge Cases"),
    ("21.7", 'Double-tap "Book" button doesn\'t create duplicate booking', "Blocker", "2", "Edge Cases"),
    ("21.8", 'Double-tap "Pay" doesn\'t charge twice', "Blocker", "2", "Edge Cases"),
    ("21.9", "Expired token mid-session: auto-refresh or re-login prompt", "Blocker", "2", "Edge Cases"),
    ("21.10", "Booking for a place that becomes full between page load and submit", "Important", "2", "Edge Cases"),
    ("21.11", "Kill app during payment flow — check no orphan payment intents", "Important", "2", "Edge Cases"),
    ("21.12", "Rate limiting: verify blocked after too many password reset attempts", "Important", "2", "Edge Cases"),
    
    ("22. VISUAL & UX POLISH", "sub"),
    ("22.1", "App cream colour (#ECE8DB) consistent across all screens", "Nice-to-have", "2", "Visual/UX"),
    ("22.2", "App blue colour (#7BA7D8) consistent for buttons/accents", "Nice-to-have", "2", "Visual/UX"),
    ("22.3", "All images load (no broken image placeholders)", "Important", "2", "Visual/UX"),
    ("22.4", "Loading states show spinners (not blank screens)", "Important", "2", "Visual/UX"),
    ("22.5", "Error messages are user-friendly (not raw errors)", "Important", "2", "Visual/UX"),
    ("22.6", 'Empty states show helpful messages ("No bookings yet")', "Important", "2", "Visual/UX"),
    ("22.7", "Keyboard doesn't cover input fields", "Important", "2", "Visual/UX"),
    ("22.8", "Scroll works properly on all long screens", "Important", "2", "Visual/UX"),
    ("22.9", "Bottom nav bar icons and labels correct", "Important", "2", "Visual/UX"),
    ("22.10", "App icon and splash screen display correctly", "Important", "2", "Visual/UX"),
    
    ("23. SECURITY CHECKS", "sub"),
    ("23.1", "Accessing someone else's booking returns error (not data)", "Blocker", "2", "Security"),
    ("23.2", "Non-admin accessing admin endpoints returns 403", "Blocker", "2", "Security"),
    ("23.3", "Non-host accessing host endpoints returns error", "Blocker", "2", "Security"),
    ("23.4", "Expired reset password link doesn't work", "Blocker", "2", "Security"),
    ("23.5", "Can't approve own place (if somehow host = admin)", "Important", "2", "Security"),
    ("23.6", "Review photos strip GPS/EXIF metadata", "Important", "2", "Security"),
    ("23.7", "Password change requires correct current password", "Blocker", "2", "Security"),
    ("23.8", "HTTPS only on all endpoints", "Blocker", "2", "Security"),
    ("23.9", "Tokens stored securely (Keychain/Keystore, not plain text)", "Blocker", "2", "Security"),
    
    ("24. PERFORMANCE", "sub"),
    ("24.1", "App cold start time < 5 seconds", "Important", "2", "Performance"),
    ("24.2", "Map panning/zooming is smooth", "Important", "2", "Performance"),
    ("24.3", "Image gallery scrolling is smooth", "Important", "2", "Performance"),
    ("24.4", "Chat messages don't cause memory leak (leave open 5+ mins)", "Important", "2", "Performance"),
    ("24.5", "Website pages load < 3 seconds on 4G", "Important", "2", "Performance"),
    ("24.6", "No console errors on website (check Chrome DevTools)", "Important", "2", "Performance"),
]

for item in tests:
    if len(item) == 2 and item[1] is None:
        # Day section header
        add_section_row(ws, row, item[0])
        row += 1
    elif len(item) == 2 and item[1] == "sub":
        add_section_row(ws, row, item[0])
        row += 1
    elif len(item) == 2 and item[1] == "subsub":
        add_subsection_row(ws, row, item[0])
        row += 1
    else:
        notes = item[5] if len(item) > 5 else ""
        add_test_row(ws, row, item[0], item[1], item[2], item[3], item[4], notes)
        row += 1

# Add data validation for Result column
from openpyxl.worksheet.datavalidation import DataValidation
dv = DataValidation(type="list", formula1='"PASS,FAIL,PARTIAL,SKIPPED,N/A"', allow_blank=True)
dv.error = "Please select PASS, FAIL, PARTIAL, SKIPPED, or N/A"
dv.errorTitle = "Invalid Result"
ws.add_data_validation(dv)
dv.add(f'F5:F{row}')

# Freeze panes
ws.freeze_panes = 'A5'

# Auto-filter
ws.auto_filter.ref = f'A4:G{row-1}'

# ========== SHEET 2: BUG TRACKER ==========
ws2 = wb.create_sheet("Bug Tracker")
ws2.sheet_properties.tabColor = "FF0000"

bug_headers = ["Bug #", "Date Found", "Section", "Severity", "Description", "Steps to Reproduce", "Expected Result", "Actual Result", "Status", "Fixed In"]
ws2.column_dimensions['A'].width = 8
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 18
ws2.column_dimensions['D'].width = 12
ws2.column_dimensions['E'].width = 40
ws2.column_dimensions['F'].width = 40
ws2.column_dimensions['G'].width = 30
ws2.column_dimensions['H'].width = 30
ws2.column_dimensions['I'].width = 12
ws2.column_dimensions['J'].width = 12

for i, h in enumerate(bug_headers, 1):
    cell = ws2.cell(row=1, column=i, value=h)
    cell.font = header_font
    cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Pre-fill 30 empty bug rows
for r in range(2, 32):
    ws2.cell(row=r, column=1, value=f"B{r-1}")
    for c in range(1, 11):
        ws2.cell(row=r, column=c).border = thin_border
        ws2.cell(row=r, column=c).alignment = wrap

# Severity dropdown
sev_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low,Cosmetic"', allow_blank=True)
ws2.add_data_validation(sev_dv)
sev_dv.add('D2:D31')

# Status dropdown
stat_dv = DataValidation(type="list", formula1='"New,In Progress,Fixed,Won\'t Fix,Deferred"', allow_blank=True)
ws2.add_data_validation(stat_dv)
stat_dv.add('I2:I31')

ws2.freeze_panes = 'A2'

# ========== SHEET 3: SUMMARY ==========
ws3 = wb.create_sheet("Summary")
ws3.sheet_properties.tabColor = "00B050"

ws3.column_dimensions['A'].width = 25
ws3.column_dimensions['B'].width = 15
ws3.column_dimensions['C'].width = 15
ws3.column_dimensions['D'].width = 15

ws3.merge_cells('A1:D1')
ws3.cell(row=1, column=1, value="Proper Place — Test Summary").font = Font(bold=True, size=14, color="2F5496")

info = [
    ("Version", "2.2.0"),
    ("Test Date", ""),
    ("Tester", "Pierce Shapton"),
    ("Platforms", "iOS, Android, Web"),
    ("Backend URL", "https://octopus-app-lxh2t.ondigitalocean.app"),
    ("Website", "proper-place.co.uk"),
]
for i, (k,v) in enumerate(info, 3):
    ws3.cell(row=i, column=1, value=k).font = Font(bold=True)
    ws3.cell(row=i, column=2, value=v)

r = 10
headers3 = ["Category", "Total", "Passed", "Failed"]
for i, h in enumerate(headers3, 1):
    cell = ws3.cell(row=r, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

categories = [
    "Blocker (🔴)", "Important (🟡)", "Nice-to-have (🟢)", "TOTAL"
]
counts = [85, 100, 45, 230]
for i, (cat, count) in enumerate(zip(categories, counts), r+1):
    ws3.cell(row=i, column=1, value=cat).border = thin_border
    ws3.cell(row=i, column=2, value=count).border = thin_border
    ws3.cell(row=i, column=3, value="").border = thin_border
    ws3.cell(row=i, column=4, value="").border = thin_border
    if cat == "TOTAL":
        ws3.cell(row=i, column=1).font = Font(bold=True)

r2 = r + len(categories) + 2
ws3.cell(row=r2, column=1, value="Known Issues").font = Font(bold=True, size=12, color="C00000")
ws3.cell(row=r2+1, column=1, value="1. Production SMTP emails failing from DigitalOcean IPs (relay propagation pending)")
ws3.cell(row=r2+2, column=1, value='2. "host_applications" table may not exist in production DB')

r3 = r2 + 5
ws3.merge_cells(f'A{r3}:D{r3}')
ws3.cell(row=r3, column=1, value="SIGN-OFF").font = Font(bold=True, size=12, color="2F5496")
signoff = [("Tester", "", ""), ("Developer", "", ""), ("Release Approved", "", "")]
headers_s = ["Role", "Name", "Date"]
for i, h in enumerate(headers_s, 1):
    cell = ws3.cell(row=r3+1, column=i, value=h)
    cell.font = Font(bold=True)
    cell.border = thin_border
for i, (role, name, date) in enumerate(signoff, r3+2):
    ws3.cell(row=i, column=1, value=role).border = thin_border
    ws3.cell(row=i, column=2, value=name).border = thin_border
    ws3.cell(row=i, column=3, value=date).border = thin_border

# ========== SHEET 4: STRIPE TEST CARDS ==========
ws4 = wb.create_sheet("Stripe Test Cards")
ws4.sheet_properties.tabColor = "635BFF"

ws4.column_dimensions['A'].width = 25
ws4.column_dimensions['B'].width = 25
ws4.column_dimensions['C'].width = 15
ws4.column_dimensions['D'].width = 10
ws4.column_dimensions['E'].width = 35

ws4.cell(row=1, column=1, value="Stripe Test Cards Reference").font = Font(bold=True, size=14, color="635BFF")

card_headers = ["Card Number", "Description", "Expiry", "CVC", "Expected Result"]
for i, h in enumerate(card_headers, 1):
    cell = ws4.cell(row=3, column=i, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="635BFF", end_color="635BFF", fill_type="solid")
    cell.border = thin_border

cards = [
    ("4242 4242 4242 4242", "Successful payment", "Any future", "Any", "Payment succeeds"),
    ("4000 0000 0000 0002", "Declined card", "Any future", "Any", "Card declined error"),
    ("4000 0025 0000 3155", "3D Secure required", "Any future", "Any", "Authentication challenge shown"),
    ("4000 0000 0000 9995", "Insufficient funds", "Any future", "Any", "Insufficient funds error"),
    ("4000 0000 0000 0069", "Expired card", "Any future", "Any", "Expired card error"),
]
for i, (num, desc, exp, cvc, result) in enumerate(cards, 4):
    ws4.cell(row=i, column=1, value=num).border = thin_border
    ws4.cell(row=i, column=1).font = Font(name='Courier New', size=11)
    ws4.cell(row=i, column=2, value=desc).border = thin_border
    ws4.cell(row=i, column=3, value=exp).border = thin_border
    ws4.cell(row=i, column=4, value=cvc).border = thin_border
    ws4.cell(row=i, column=5, value=result).border = thin_border

ws4.cell(row=10, column=1, value="Admin Login").font = Font(bold=True, size=12)
ws4.cell(row=11, column=1, value="Email:").font = Font(bold=True)
ws4.cell(row=11, column=2, value="admin@properplace.com")
ws4.cell(row=12, column=1, value="Password:").font = Font(bold=True)
ws4.cell(row=12, column=2, value="AdminPass123!")

# Save
output = "/Users/pierceshaptonproperplace/Proper-Place-V2/APP_RELEASE_TEST_PLAN.xlsx"
wb.save(output)
print(f"✅ Excel test plan saved to: {output}")
print(f"   Sheets: {wb.sheetnames}")
