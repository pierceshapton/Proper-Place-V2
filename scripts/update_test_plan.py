import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from copy import copy

INPUT = '/Users/pierceshaptonproperplace/Downloads/APP_RELEASE_TEST_PLAN (1).xlsx'
OUTPUT = '/Users/pierceshaptonproperplace/Downloads/APP_RELEASE_TEST_PLAN_UPDATED.xlsx'

wb = openpyxl.load_workbook(INPUT)
ws = wb['Test Plan']

# Map of test IDs -> (new_status, notes_update)
fixes = {
    # FLUTTER FIXES
    '1.6.4': ('FIXED — RETEST', 'Fixed: App now shows cream loading screen while checking auth, skips welcome if token exists.'),
    '2.1.7': ('FIXED — RETEST', 'Fixed: Map no longer fetches GPS on tab switch. Uses cached position from last camera move. Only gets GPS on first-ever launch.'),
    '2.2.6': ('NOTE', 'Search already searched addresses. Now also confirmed in code — searches both name and address.'),
    '3.6':   ('FIXED — RETEST', 'Fixed: Stars now use actual DB rating (0 if no reviews). Hidden entirely when rating is 0. No more hardcoded 4.5.'),
    '3.12':  ('NOTE', 'Investigated: Host name does NOT display in the code. Only place/site name appears. No code change needed.'),
    '3.14':  ('FIXED — RETEST', 'Fixed: Removed "added/removed from favourites" snackbar. Heart icon toggle is the only feedback.'),
    '4.1.2': ('FIXED — RETEST', 'Fixed: Calendar booked-date colour changed from green (#81C784) to app blue (#7BA7D8).'),
    '4.3.4': ('FIXED — RETEST', 'Fixed: Booking search now also matches place_address/address field in addition to name.'),
    '4.3.9': ('FIXED — RETEST', 'Fixed: Cancel booking request now includes Authorization header with Bearer token.'),
    '4.3.11':('FIXED — RETEST', 'Fixed: Booking overlap check changed to strict (< >) so checkout day can be another booking\'s check-in day.'),
    '4.3.12':('FIXED — RETEST', 'Fixed: HomeScreen.build() now checks for pending tab/focus place set by static methods, applies them on popUntil return.'),
    '7.1':   ('FIXED — RETEST', 'Fixed: Added heart/favourites filter button on map screen (below filter button). Toggles to show only favourited places.'),
    '8.1.5': ('FIXED — RETEST', 'Fixed: Address search popup height changed from 75% to 85% of screen.'),
    '13.1':  ('FIXED — RETEST', 'Fixed: All labels, section headers, and text field values on profile screen now explicitly set to black.'),
    '17.1':  ('FIXED — RETEST', 'Fixed: Price keypad Done/Cancel toolbar now positioned at bottom:0 (was double-offset by keyboard height).'),
    
    # BACKEND FIXES
    '12.4.2':('FIXED — RETEST', 'Fixed: Admin users now bypass 72-hour chat window restriction. Can message anytime.'),
    '15.1.11':('FIXED — RETEST', 'Fixed: Contact form userId made optional — no longer requires logged-in user ID.'),
    '15.4.5':('FIXED — RETEST', 'Fixed: Chat sendMessage now accepts both camelCase (receiverId) and snake_case (receiver_id) params.'),
    '20.1':  ('FIXED — RETEST', 'Fixed: Referral code generation now requires host role AND at least one approved site.'),
    
    # WEBSITE FIXES
    '15.3.2':('FIXED — RETEST', 'Fixed: Replaced <input type="date"> with visual calendar component showing unavailable dates, range selection, and colour coding.'),
    '15.4.3':('FIXED — RETEST', 'Fixed: Backend enforces 24-hour cancellation deadline. Website now shows warning modal with 24h policy instead of browser confirm().'),
    '15.4.6':('FIXED — RETEST', 'Fixed: Added useEffect to sync profile/vehicle state when user data loads from auth context.'),
    '15.4.11':('FIXED — RETEST', 'Fixed: Added "Host Earnings" card to dashboard showing total from confirmed+completed bookings.'),
    '15.5.3':('FIXED — RETEST', 'Fixed: Replaced browser prompt() with in-website ReasonModal for all rejection flows (admin places, host bookings, host applications).'),
    '15.5.4':('FIXED — RETEST', 'Fixed: Admin bookings table now shows booking_ref instead of database ID.'),
    '17.5':  ('FIXED — RETEST', 'Fixed: Admin places summary now shows description preview, photo count, amenity count, and vehicle limit indicators.'),
    '22.3':  ('FIXED — RETEST', 'Same as 17.5 — admin site details now show more info in summary row.'),
    
    # UNCHANGED / INVESTIGATED / SKIPPED
    '1.3.2': ('SKIPPED', 'SMTP blocked from DigitalOcean production IPs. Email sending requires migration to SendGrid/SES.'),
    '1.3.5': ('SKIPPED', 'SMTP blocked — same as 1.3.2.'),
    '4.3.0': ('NOTE', 'Payment error for different location is related to user already having active booking. Calendar crossover is by design (one site per day). Could add clearer UX messaging.'),
    '4.3.10':('NOTE', 'Depends on 4.3.9 fix. Retest cancel flow after auth fix.'),
    '5.2.2': ('NOTE', 'Chat "closing soon" indicator — requires frontend countdown timer UI. Not yet implemented.'),
    '8.1.10':('NOTE', 'Host applications endpoint exists and is properly routed (POST /host-applications). Previous error may have been transient deployment issue. Retest.'),
    '8.2.1': ('NOTE', 'Host contract screen depends on 8.1.10 working. Retest after confirming host application flow.'),
    '13.6':  ('NOTE', 'Vehicle dimensions code looks functional — sliders and toggles have proper handlers. May need device retest.'),
    '18.4':  ('NOTE', 'Stripe fee working as designed. 15% Proper Place fee only applies when host has Stripe Connect account (destination charges). Test hosts don\'t have Connect yet.'),
    '19.4':  ('NOTE', 'Admin contact messages section exists. May need retest after contact form fix (15.1.11).'),
    '19.5':  ('NOTE', 'Contact status management exists in admin routes. Retest.'),
    '2.3.1': ('PARTIAL', 'Height fixed to 85%. Route calculation not yet implemented — needs Google Directions API integration & polyline drawing.'),
    '2.1.2': ('NOTE', 'On simulator, GPS returns San Francisco coords which are deliberately skipped. Works on real device.'),
    '15.2.4':('SKIPPED', 'Depends on SMTP — same as 1.3.2.'),
    '4.3.7': ('NOTE', 'Waze link uses generic lat/lng URL. May need Waze-specific deep link format for destination. Google Maps works fine.'),
    '15.5.5':('NOTE', 'Admin support ticket section exists. Open/filter behaviour and ticket detail view need further UX work.'),
    '22.7':  ('NOTE', 'Keyboard coverage is context-specific. Price keypad toolbar positioning fixed (17.1).'),
    '1.2.6': ('NOTE', 'No notes provided on original test. Needs clarification.'),
}

# Green fill for FIXED, yellow for NOTE, light blue for SKIPPED
fill_fixed = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
fill_note = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
fill_skipped = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
font_fixed = Font(color='006100', bold=True)
font_note = Font(color='9C6500')
font_skipped = Font(color='305496')

updated = 0
for r in range(1, ws.max_row + 1):
    test_id = str(ws.cell(r, 1).value or '').strip()
    if test_id in fixes:
        new_status, new_notes = fixes[test_id]
        
        # Update Result column (6)
        cell_result = ws.cell(r, 6)
        cell_result.value = new_status
        
        # Update Notes column (7)
        cell_notes = ws.cell(r, 7)
        old_notes = cell_notes.value or ''
        cell_notes.value = f'{new_notes}\n\n[Original: {old_notes}]' if old_notes else new_notes
        cell_notes.alignment = Alignment(wrap_text=True)
        
        # Colour coding
        if 'FIXED' in new_status:
            cell_result.fill = fill_fixed
            cell_result.font = font_fixed
        elif new_status == 'NOTE':
            cell_result.fill = fill_note
            cell_result.font = font_note
        elif new_status == 'SKIPPED':
            cell_result.fill = fill_skipped
            cell_result.font = font_skipped
        
        updated += 1

# Update Bug Tracker sheet too
ws_bugs = wb['Bug Tracker']
# Add fix summary rows
next_row = ws_bugs.max_row + 2
ws_bugs.cell(next_row, 1, value='--- FIXES APPLIED ---')
ws_bugs.cell(next_row, 1).font = Font(bold=True)
next_row += 1

fix_descriptions = [
    ('3.6', 'Hardcoded 4.5 stars', 'Stars now from DB, hidden if 0'),
    ('2.1.7', 'Map resets position on tab switch', 'Cached position used, no GPS re-fetch'),
    ('1.6.4', 'Welcome screen flash on resume', 'Loading screen shown while checking auth'),
    ('4.3.9', 'Cancel booking auth error', 'Added Bearer token to cancel request'),
    ('4.3.11', 'Can\'t book on checkout day', 'Strict overlap comparison allows same-day transitions'),
    ('4.1.2', 'Calendar green instead of blue', 'Changed to app blue #7BA7D8'),
    ('3.14', 'Favourites snackbar unnecessary', 'Removed snackbar, heart icon only'),
    ('13.1', 'Profile text not black', 'All labels and text explicitly black'),
    ('12.4.2', 'Admin messaging 403', 'Admin bypasses 72h chat window'),
    ('15.1.11', 'Contact form "missing fields"', 'userId made optional'),
    ('15.4.5', 'Chat "receiverId required"', 'Accepts both camelCase and snake_case'),
    ('15.4.6', 'Profile shows defaults', 'useEffect syncs state on load'),
    ('15.4.3', 'Cancel in restricted period', '24h deadline enforced, warning modal'),
    ('20.1', 'Referral not host-gated', 'Requires host role + approved site'),
    ('15.3.2', 'Plain date picker', 'Visual calendar with unavailable dates'),
    ('15.5.3', 'Chrome prompt() popups', 'In-website ReasonModal component'),
    ('15.5.4', 'Random booking IDs', 'Shows booking_ref'),
    ('17.5', 'Admin site lacks details', 'Added description, photos, amenities count'),
    ('7.1', 'No favourites on map', 'Added heart filter button'),
    ('4.3.4', 'Booking search missing address', 'Added address field to search'),
    ('4.3.12', 'Show on map returns to bookings', 'Fixed tab switch on popUntil'),
    ('8.1.5', 'Popup too short', 'Changed to 85% height'),
    ('17.1', 'Keypad toolbar too high', 'Fixed bottom:0 positioning'),
    ('15.4.11', 'No host earnings section', 'Added Host Earnings card to dashboard'),
]

for test_id, desc, fix in fix_descriptions:
    ws_bugs.cell(next_row, 1, value=test_id)
    ws_bugs.cell(next_row, 5, value=desc)
    ws_bugs.cell(next_row, 8, value=fix)
    ws_bugs.cell(next_row, 9, value='FIXED')
    ws_bugs.cell(next_row, 9).fill = fill_fixed
    ws_bugs.cell(next_row, 9).font = font_fixed
    next_row += 1

wb.save(OUTPUT)
print(f'Updated {updated} test items. Saved to {OUTPUT}')
